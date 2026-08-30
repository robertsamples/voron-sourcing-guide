"""Read the published guide's row colour coding so the rebuild can match it.

The guide fills each row with a colour for its category. Tabs disagree in
places - VORON 0.2 uses a different green for Electronics than 2.4 does - so
the newest tab that defines a category wins, and the older ones are reported
rather than silently dropped.

Writes data/category_fills.json, keyed by the master's normalised category
names so tools/build_xlsx.py can look colours up directly.

    python tools/extract_styles.py
"""
import json
import os
import sys
from collections import Counter, defaultdict

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from unify import norm_category  # noqa: E402

XLSX = "Published Voron Development Team Sourcing Guide.xlsx"
OUT = "data/category_fills.json"

# newest first: a category's colour is taken from the most current tab that
# uses it, so retired tabs cannot override a shipping one
PRECEDENCE = ["VORON 2.4", "VORON Trident", "VORON 0.2", "Voron Switchwire",
              "Voron Stealthburner", "VORON Optional Parts", "VORON Tools"]


def main():
    wb = openpyxl.load_workbook(XLSX)
    per_sheet = {}
    for name in PRECEDENCE:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        seen = defaultdict(Counter)
        for row in ws.iter_rows(min_row=2):
            cat = row[0].value
            if not isinstance(cat, str) or not cat.strip():
                continue
            fill = row[0].fill
            if fill and fill.fill_type and fill.fgColor.rgb:
                seen[norm_category(cat)][fill.fgColor.rgb] += 1
        per_sheet[name] = {k: v.most_common(1)[0][0] for k, v in seen.items()}

    merged, source, clashes = {}, {}, {}
    for name in PRECEDENCE:
        for cat, rgb in per_sheet.get(name, {}).items():
            if cat not in merged:
                merged[cat], source[cat] = rgb, name
            elif merged[cat] != rgb:
                clashes.setdefault(cat, []).append("%s=%s" % (name, rgb))

    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(dict(sorted(merged.items())), fh, indent=1)

    print("%d categories\n" % len(merged))
    for cat in sorted(merged):
        note = ""
        if cat in clashes:
            note = "  (older tabs differ: %s)" % ", ".join(clashes[cat])
        print("  %-24s %s  from %s%s" % (cat, merged[cat], source[cat], note))
    print("\nwrote %s" % OUT)


if __name__ == "__main__":
    main()
