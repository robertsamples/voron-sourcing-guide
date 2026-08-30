"""Stage 1: flatten every sourcing-guide tab into one record per spreadsheet row.

Nothing is merged or renamed here beyond decoding links -- this file is the
audit trail that every later stage is derived from.
"""
import json
import os
import re
import openpyxl

os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

XLSX = "Published Voron Development Team Sourcing Guide.xlsx"
OUT = "data/raw_extract.json"

META_HEADERS = {
    "": "category", " ": "category", "category": "category",
    "component": "component",
    "standard": "standard",
    "qty": "qty",
    "size": "size", "all sizes": "size",
}
NOTES_HEADERS = {"notes"}

# A HYPERLINK argument is one or more quoted strings joined by & -- a handful of
# rows in the workbook carry AliExpress URLs split that way.
_ARG = r'"(?:[^"]|"")*"(?:\s*&\s*"(?:[^"]|"")*")*'
HYPERLINK_RE = re.compile(
    r'^\s*=\s*HYPERLINK\(\s*(%s)\s*(?:,\s*(%s)\s*)?\)\s*$' % (_ARG, _ARG),
    re.IGNORECASE,
)


def unesc(arg):
    """Join a concatenated quoted-string formula argument into plain text."""
    if arg is None:
        return None
    return "".join(m.group(1).replace('""', '"')
                   for m in re.finditer(r'"((?:[^"]|"")*)"', arg))


def norm_text(v):
    if v is None:
        return None
    if not isinstance(v, str):
        return v
    s = v.replace(" ", " ").strip()
    s = re.sub(r"[ \t]+", " ", s)
    return s or None


def cell_link(cell, vcell=None):
    """Return (text, url), decoding =HYPERLINK() formulas and real hyperlinks.

    vcell is the same cell from the data_only workbook; it supplies the cached
    result for any other formula (a few tabs compute Misumi part numbers).
    """
    v = cell.value
    url = text = None
    if isinstance(v, str):
        m = HYPERLINK_RE.match(v)
        if m:
            url = unesc(m.group(1))
            text = unesc(m.group(2)) if m.group(2) is not None else url
        else:
            text = v
            if v.lstrip().startswith("=") and vcell is not None:
                cached = norm_text(vcell.value)
                if cached is not None:
                    text = cached
    elif v is not None:
        text = v
    if url is None and cell.hyperlink is not None and cell.hyperlink.target:
        url = cell.hyperlink.target
    return norm_text(text), url


# some tabs label the size column with a build volume, e.g. "230³"
SIZE_HEADER_RE = re.compile(r"^\d{2,4}\s*[³°]?$")


def classify(label):
    key = (label or "").strip().lower()
    if SIZE_HEADER_RE.match(key):
        return "size"
    if key in META_HEADERS:
        return META_HEADERS[key]
    if key in NOTES_HEADERS:
        return "notes"
    return "vendor"


def build_layout(ws):
    """Map every column -> slot index. Slots absorb unlabelled columns."""
    anchors = []
    for cell in ws[1]:
        lab = norm_text(cell.value)
        if lab is None and cell.column != 1:
            continue
        anchors.append((cell.column, classify(lab), lab if lab is not None else "Category"))
    anchors.sort()
    if not anchors:
        return {}, []

    max_col = ws.max_column
    slots = []
    for i, (col, kind, lab) in enumerate(anchors):
        nxt = anchors[i + 1][0] if i + 1 < len(anchors) else max_col + 1
        slots.append({"name": lab, "kind": kind, "start": col, "end": nxt - 1})

    owner = {}
    for i, s in enumerate(slots):
        cols = list(range(s["start"], s["end"] + 1))
        if s["kind"] in ("vendor", "notes"):
            for c in cols:
                owner[c] = i
        else:
            # A metadata header owns only its own column. Unlabelled columns to
            # its right belong to the next slot -- in several tabs the vendor
            # header sits one column right of where its data actually starts.
            owner[cols[0]] = i
            nxt_i = i + 1 if i + 1 < len(slots) else i
            for c in cols[1:]:
                owner[c] = nxt_i
    return owner, slots


def merge_map(ws):
    m = {}
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                m[(r, c)] = (rng.min_row, rng.min_col, rng.max_col)
    return m


def main():
    wb = openpyxl.load_workbook(XLSX)                    # formulas
    wbv = openpyxl.load_workbook(XLSX, data_only=True)   # cached values
    rows_out = []
    layouts = {}

    for ws in wb.worksheets:
        wsv = wbv[ws.title]
        owner, slots = build_layout(ws)
        layouts[ws.title] = slots
        if not slots:
            continue
        mm = merge_map(ws)

        def first(kind):
            return next((s["start"] for s in slots if s["kind"] == kind), None)

        cat_col, comp_col = first("category"), first("component")
        std_col, qty_col, size_col = first("standard"), first("qty"), first("size")
        meta_cols = {c for c in (cat_col, comp_col, std_col, qty_col, size_col) if c}

        last_cat = last_comp = last_std = None
        for row in range(2, ws.max_row + 1):
            def val(col):
                if col is None:
                    return None
                a = mm.get((row, col))
                if a:
                    # a merge that starts in a *different* metadata column (e.g.
                    # Component merged across B:C) must not leak into this one
                    if a[1] != col and a[1] in meta_cols:
                        return None
                    return norm_text(wsv.cell(row=a[0], column=a[1]).value)
                return norm_text(wsv.cell(row=row, column=col).value)

            cat, comp = val(cat_col), val(comp_col)
            std, qty, size = val(std_col), val(qty_col), val(size_col)

            entries = []
            for col in range(1, ws.max_column + 1):
                if col in meta_cols:
                    continue
                oi = owner.get(col)
                if oi is None:
                    continue
                anchor = mm.get((row, col))
                if anchor and anchor[1] != col:
                    continue  # read each merged block once, at its anchor
                text, url = cell_link(ws.cell(row=row, column=col),
                                      wsv.cell(row=row, column=col))
                if text is None and url is None:
                    continue
                slot_i = oi
                if anchor:
                    # a merged block reaching into a vendor slot belongs there
                    for c2 in range(anchor[1], anchor[2] + 1):
                        o2 = owner.get(c2)
                        if o2 is not None and slots[o2]["kind"] == "vendor":
                            slot_i = o2
                            break
                entries.append({
                    "col": ws.cell(row=row, column=col).column_letter,
                    "slot": slots[slot_i]["name"],
                    "kind": slots[slot_i]["kind"],
                    "text": text,
                    "url": url,
                })

            if comp is None and cat is None and not entries:
                continue

            # A footer line is not a component, so a blank row under it is not
            # a continuation of anything. VORON 1.8 keeps a frame calculator
            # below its affiliate footer, and without this those rows inherit
            # the footer's text as their part name.
            is_footer = bool(comp and re.match(r"^\W*as an (amazon|aliexpress)",
                                               comp, re.I))
            is_cont = comp is None
            if is_cont:
                comp, cat, std = last_comp, cat or last_cat, std or last_std
            elif not is_footer:
                last_comp, last_cat, last_std = comp, cat, std
            else:
                last_comp = last_cat = last_std = None
            if comp is None:
                continue

            rows_out.append({
                "sheet": ws.title,
                "row": row,
                "continuation": is_cont,
                "category": cat,
                "component": comp,
                "standard": std,
                "qty": qty,
                "size": size,
                "notes": [e["text"] for e in entries if e["kind"] == "notes" and e["text"]],
                "entries": [e for e in entries if e["kind"] == "vendor"],
            })

    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump({"layouts": layouts, "rows": rows_out}, fh,
                  indent=1, ensure_ascii=False, default=str)
    print("rows:", len(rows_out),
          "vendor cells:", sum(len(r["entries"]) for r in rows_out),
          "with urls:", sum(1 for r in rows_out for e in r["entries"] if e["url"]))


if __name__ == "__main__":
    main()
