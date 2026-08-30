"""Write a revised sourcing workbook from the master.

Same shape as the published guide - one sheet per tab, Category / Component /
Standard / Qty / Size, then vendor columns - with three differences:

  * component names are the canonical ones, so a part is spelled the same way
    on every tab it appears on
  * every live link for a component is offered, not just the ones that happened
    to be typed on that tab
  * dead links are gone, and a final column says what was removed from the row

Nothing that is not `link_ok == "yes"` is written. A row whose links were all
dead keeps its quantities and notes but has no vendor - that is the honest
result, and the last column says so.

    python tools/build_xlsx.py
"""
import json
import os
import re
import sys
from collections import OrderedDict

from normalize import clean_text, is_affiliate_label, name_key
from unify import PROJECTS, is_drop_row, load_aliases

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

MASTER = "data/voron_sourcing_master.json"
OUT = "Voron Sourcing Guide (revised).xlsx"

META = ["Category", "Component", "Standard", "Qty", "Size"]
ROLE_RANK = {"recommended": 0, "alternative": 1, "budget": 2,
             "non_affiliate": 3, "prusa_salvage": 4, "unknown": 5}

# Styling copied from the published guide. Where tabs disagree the newest one
# wins, so these come from VORON 2.4 first, then Trident, 0.2, Switchwire,
# Stealthburner (see tools/extract_styles.py).
HEAD_FILL = PatternFill("solid", fgColor="FFD9D9D9")
HEAD_FONT = Font(bold=True, size=11)
LINK_FONT = Font(color="FF1155CC", underline="single")
REC_FONT = Font(color="FF1155CC", underline="single", bold=True)
WARN_FONT = Font(color="FF9B2C24")
NOTE_ALIGN = Alignment(wrap_text=True, vertical="top")

CATEGORY_FILLS = json.load(open("data/category_fills.json", encoding="utf-8"))

# widths follow VORON 2.4; the guide's Notes column is enormous, trimmed to
# something that still fits on a screen
WIDTHS = {"category": 19.9, "component": 34.5, "standard": 17.1,
          "qty": 7.0, "size": 8.5, "vendor": 20.0, "affiliate": 15.0,
          "notes": 90.0, "check": 20.0}


def slot_names(n):
    """Column headings for n vendor slots, in the guide's own vocabulary."""
    out = []
    for i in range(n):
        if i == 0:
            out.append("Recommended")
        elif i == 1:
            out.append("Alternative Source")
        else:
            out.append("Alternative Source %d" % i)
    return out


def live_sources(item):
    """Live links for a component, best first."""
    srcs = [s for s in item["sources"] if s["link_ok"] == "yes"]
    srcs.sort(key=lambda s: (min(ROLE_RANK.get(r, 9) for r in s["roles"]),
                             -len(s["projects"]), s["vendor_name"] or ""))
    return srcs


def label_for(src):
    """Prefer the wording the guide already used, but never an affiliate tag."""
    for lab in src["labels"]:
        if lab and not is_affiliate_label(lab):
            return lab
    return src["vendor_name"] or src["url"]


def text_sources(item, project):
    """Guidance the guide gave instead of a link, for this tab.

    "Print Yourself", "included in kit", "cut by hand per Drawings" - not links,
    so nothing to check, but dropping them would lose the instruction.
    """
    return [u["text"] for u in item["unlinked_sources"]
            if project in u["projects"]]


def live_affiliate(src):
    for url, ok in zip(src["affiliate_urls"], src["affiliate_link_ok"]):
        if ok == "yes":
            return url
    return None


def removal_note(item, project):
    """What this tab lost, judged against the links this tab used to carry."""
    had = [s for s in item["sources"]
           if any(v["project"] == project for v in s["seen"])]
    if not had:
        return ""
    dead = [s for s in had if s["link_ok"] != "yes"]
    if not dead:
        return ""
    if len(dead) == len(had):
        return "all links were dead"
    return "some links were dead"


def structural_rows(sheet, dropped):
    """The rows that organise a tab rather than list a part.

    `OR` separators between alternatives, headings like "Full Printer kits are
    availible from a number of trusted suppliers listed below", and the
    affiliate disclosure at the foot. They carry no link and are not components,
    but dropping them would change what the tab means.
    """
    raw = json.load(open("data/raw_extract.json", encoding="utf-8"))
    out = []
    for r in raw["rows"]:
        if r["sheet"] != sheet or r["continuation"]:
            continue
        comp = clean_text(r["component"])
        if not comp:
            continue
        key = name_key(comp)
        if not (is_drop_row(comp) or key in dropped):
            continue
        out.append({
            "order": r["row"],
            "category": clean_text(r["category"]),
            "text": comp,
            "kind": "or" if comp.strip().upper() == "OR" else (
                "footer" if key.startswith("as an ") else "heading"),
        })
    return out


def rows_for(project, items, sheet, dropped, master):
    """Every row of the tab, parts and structure alike, in its own order."""
    rows = []
    for item in items:
        use = next((u for u in item["used_by"] if u["project"] == project), None)
        if use is None:
            continue
        order = min(use["rows"]) if use["rows"] else 10 ** 6
        sizes = use["qty"] or [{"size": None, "qty": None}]
        for n, q in enumerate(sizes):
            rows.append((order, n, "part", (item, use, q)))
    for st in structural_rows(sheet, dropped):
        rows.append((st["order"], 0, "structure", st))
    rows.sort(key=lambda r: (r[0], r[1]))
    return drop_stranded_ors(rows, neighbour_ids(sheet, master))


def neighbour_ids(sheet, master):
    """For each `OR` row, the component ids of the rows it sits between.

    An OR separates two alternatives in the original sheet. After
    canonicalising, both sides may now be the same component - the BMG clone,
    its insides and the drive gear kit merged into one - and the separator has
    nothing left to separate.
    """
    by_key = {}
    for item in master["items"]:
        for n in [item["name"]] + item["aliases"]:
            by_key.setdefault(name_key(n), item["id"])
    raw = json.load(open("data/raw_extract.json", encoding="utf-8"))
    seq = [r for r in raw["rows"] if r["sheet"] == sheet and not r["continuation"]]
    seq.sort(key=lambda r: r["row"])
    out = {}
    for n, r in enumerate(seq):
        if (clean_text(r["component"]) or "").strip().upper() != "OR":
            continue
        def side(rng):
            for i in rng:
                comp = clean_text(seq[i]["component"])
                if comp and comp.strip().upper() != "OR":
                    return by_key.get(name_key(comp))
            return None
        out[r["row"]] = (side(range(n - 1, -1, -1)), side(range(n + 1, len(seq))))
    return out


def drop_stranded_ors(rows, or_sides):
    """Remove `OR` rows that no longer separate two different components.

    Canonicalising merged some of the alternatives an OR used to sit between -
    the BMG clone, its insides and the drive gear kit are one component now -
    which would otherwise leave the separator dangling between a row and
    itself, or two of them back to back.
    """
    out = []
    for n, row in enumerate(rows):
        if row[2] == "structure" and row[3]["kind"] == "or":
            before, after = or_sides.get(row[0], (None, None))
            if before is None or after is None or before == after:
                continue
            if out and out[-1][2] == "structure" and out[-1][3]["kind"] == "or":
                continue
        out.append(row)
    return out


def main():
    m = json.load(open(MASTER, encoding="utf-8"))
    _, _, dropped, _ = load_aliases()
    wb = Workbook()
    wb.remove(wb.active)

    by_project = OrderedDict((p["id"], []) for p in m["projects"])
    for item in m["items"]:
        for u in item["used_by"]:
            if u["project"] in by_project:
                by_project[u["project"]].append(item)

    summary = []
    for p in m["projects"]:
        items = by_project[p["id"]]
        if not items:
            continue
        rows = rows_for(p["id"], items, p["sheet"], dropped, m)

        widest = max((len(live_sources(i)) + len(text_sources(i, p["id"]))
                      for i in items), default=0)
        any_aff = [False] * widest
        for item in items:
            for n, s in enumerate(live_sources(item)):
                if live_affiliate(s):
                    any_aff[n] = True

        ws = wb.create_sheet(p["name"][:31])
        header, colmap, kinds = list(META), {}, ["category", "component",
                                                "standard", "qty", "size"]
        for n, name in enumerate(slot_names(widest)):
            colmap[n] = len(header) + 1
            header.append(name)
            kinds.append("vendor")
            if any_aff[n]:
                header.append("")
                kinds.append("affiliate")
        header += ["Notes", "Link check"]
        kinds += ["notes", "check"]
        ws.append(header)
        for c, kind in enumerate(kinds, start=1):
            cell = ws.cell(row=1, column=c)
            cell.fill, cell.font = HEAD_FILL, HEAD_FONT
            cell.alignment = Alignment(
                horizontal="right" if kind == "vendor" else "center",
                vertical="center")

        stripped = kept = 0
        last_fill = None
        for _, _, kind, payload in rows:
            r = ws.max_row + 1

            if kind == "structure":
                st = payload
                if st["category"]:
                    ws.cell(row=r, column=1, value=st["category"])
                cell = ws.cell(row=r, column=2, value=st["text"])
                cell.font = Font(bold=st["kind"] != "footer")
                if st["kind"] == "or":
                    cell.alignment = Alignment(horizontal="center")
                rgb = CATEGORY_FILLS.get(st["category"] or "") or (
                    last_fill if st["kind"] == "or" else None)
                if rgb:
                    band = PatternFill("solid", fgColor=rgb)
                    for c in range(1, len(header) + 1):
                        ws.cell(row=r, column=c).fill = band
                continue

            item, use, q = payload
            ws.cell(row=r, column=1, value=item["category"])
            ws.cell(row=r, column=2, value=item["name"])
            ws.cell(row=r, column=3, value=item["standard"])
            ws.cell(row=r, column=4, value=q.get("qty"))
            ws.cell(row=r, column=5, value=q.get("size"))

            for n, s in enumerate(live_sources(item)):
                col = colmap[n]
                cell = ws.cell(row=r, column=col)
                cell.value = '=HYPERLINK("%s","%s")' % (
                    s["url"].replace('"', '""'), label_for(s).replace('"', '""'))
                cell.font = REC_FONT if n == 0 else LINK_FONT
                aff = live_affiliate(s)
                if any_aff[n] and aff:
                    a = ws.cell(row=r, column=col + 1)
                    a.value = '=HYPERLINK("%s","[Affiliate Link]")' % \
                        aff.replace('"', '""')
                    a.font = LINK_FONT

            base = len(live_sources(item))
            for n, txt in enumerate(text_sources(item, p["id"])):
                if base + n in colmap:
                    ws.cell(row=r, column=colmap[base + n], value=txt)

            note_col = len(header) - 1
            notes = " / ".join(n["text"] for n in item["notes"]
                               if p["id"] in n["projects"])
            if notes:
                nc = ws.cell(row=r, column=note_col, value=notes)
                nc.alignment = NOTE_ALIGN
            warn = removal_note(item, p["id"])
            if warn:
                wc = ws.cell(row=r, column=note_col + 1, value=warn)
                wc.font = WARN_FONT
                stripped += 1
            if live_sources(item) or text_sources(item, p["id"]):
                kept += 1

            # the guide colour-codes each row by its category
            rgb = CATEGORY_FILLS.get(item["category"] or "")
            if rgb:
                last_fill = rgb
                band = PatternFill("solid", fgColor=rgb)
                for c in range(1, len(header) + 1):
                    ws.cell(row=r, column=c).fill = band

        for n, kind in enumerate(kinds, start=1):
            ws.column_dimensions[get_column_letter(n)].width = WIDTHS[kind]
        ws.freeze_panes = "A2"
        summary.append((p["name"], p["status"],
                        sum(1 for x in rows if x[2] == "part"), widest,
                        len(rows) - kept, stripped))

    ws = wb.create_sheet("About", 0)
    ws.append(["Voron sourcing guide - revised"])
    ws["A1"].font = Font(bold=True, size=14)
    for line in [
        "",
        "Rebuilt from the published guide by tools/build_xlsx.py.",
        "",
        "  * component names are canonical, so a part reads the same on every tab",
        "  * every live link for a component is listed, not only the ones that "
        "happened to be typed on that tab",
        "  * links that did not resolve when checked have been removed",
        "  * the last column of each sheet says what a row lost",
        "",
        "A row with no vendor is a row whose every link was dead. It kept its "
        "quantity and notes so the part is not lost.",
        "",
    ]:
        ws.append([line])
    ws.append(["Tab", "Status", "Rows", "Vendor slots",
               "Rows with no live link", "Rows that lost links"])
    hdr = ws.max_row
    for c in range(1, 7):
        ws.cell(row=hdr, column=c).font = Font(bold=True)
    for row in summary:
        ws.append(list(row))
    ws.column_dimensions["A"].width = 26
    for c in "BCDEF":
        ws.column_dimensions[c].width = 21

    wb.save(OUT)
    print("wrote %s\n" % OUT)
    print("  %-22s %-9s %5s %6s %8s %8s"
          % ("tab", "status", "rows", "slots", "no link", "stripped"))
    for name, status, rows, slots, nolink, stripped in summary:
        print("  %-22s %-9s %5d %6d %8d %8d"
              % (name, status, rows, slots, nolink, stripped))


if __name__ == "__main__":
    main()
