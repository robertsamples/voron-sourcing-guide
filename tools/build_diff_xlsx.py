"""Write a redline workbook: the published guide against the rebuild.

Laid out on the original's rows so the two are comparable line by line. Inside
each cell the change is shown inline, word by word:

    red, struck through   text the rebuild removed
    blue, bold            text the rebuild added
    plain                 text that did not change

Word-level rather than cell-level, so dropping a trailing note strikes the note
and leaves the rest of the cell alone instead of rewriting the whole thing.

Vendor links are matched on their URL, not their column position, so a link
that merely moved slot reads as unchanged; only genuinely removed and added
links get marked.

    python tools/build_diff_xlsx.py
"""
import json
import os
import re
import sys
from collections import OrderedDict, defaultdict
from difflib import SequenceMatcher

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from normalize import (clean_text, clean_qty, clean_size, name_key,
                       url_key)  # noqa: E402
from unify import is_drop_row, load_aliases, norm_category  # noqa: E402
from build_xlsx import (label_for, live_affiliate, live_sources,
                        text_sources)  # noqa: E402

MASTER = "data/voron_sourcing_master.json"
RAW = "data/raw_extract.json"
OUT = "Voron Sourcing Guide (diff).xlsx"

DEL = InlineFont(color="FF9B2C24", strike=True)
ADD = InlineFont(color="FF1155CC", b=True)

HEAD_FILL = PatternFill("solid", fgColor="FFD9D9D9")
HEAD_FONT = Font(bold=True, size=11)
GONE_FILL = PatternFill("solid", fgColor="FFF4DEDB")
NEW_FILL = PatternFill("solid", fgColor="FFDDE7F7")
WRAP = Alignment(wrap_text=True, vertical="top")

TOKEN = re.compile(r"\S+\s*")


def tokens(text):
    return TOKEN.findall(text or "")


def diff_cell(old, new):
    """Rich text showing old -> new, or a plain string when nothing changed."""
    old, new = (old or ""), (new or "")
    if old == new:
        return old or None
    a, b = tokens(old), tokens(new)
    parts, changed = [], False
    for op, i1, i2, j1, j2 in SequenceMatcher(None, a, b).get_opcodes():
        if op == "equal":
            parts.append("".join(a[i1:i2]))
        else:
            if i1 != i2:
                parts.append(TextBlock(DEL, "".join(a[i1:i2])))
                changed = True
            if j1 != j2:
                parts.append(TextBlock(ADD, "".join(b[j1:j2])))
                changed = True
    if not changed:
        return old or None
    return CellRichText(parts)


def link_text(label, url):
    # a few vendor cells hold a bare number rather than text
    label = "" if label is None else str(label)
    return "%s  <%s>" % (label, url) if url else label


def load_original():
    """Original rows, per sheet, in sheet order."""
    raw = json.load(open(RAW, encoding="utf-8"))
    per_sheet = defaultdict(list)
    for r in raw["rows"]:
        per_sheet[r["sheet"]].append(r)
    for rows in per_sheet.values():
        rows.sort(key=lambda r: r["row"])
    return per_sheet


def main():
    m = json.load(open(MASTER, encoding="utf-8"))
    _, _, dropped, _ = load_aliases()
    originals = load_original()

    # every spelling the guide used -> the component it became
    by_key = {}
    for item in m["items"]:
        for n in [item["name"]] + item["aliases"]:
            by_key.setdefault(name_key(n), item)

    wb = Workbook()
    wb.remove(wb.active)
    summary = []

    for p in m["projects"]:
        rows = originals.get(p["sheet"], [])
        if not rows:
            continue

        # how many link columns the widest row needs, either side
        widest = 1
        for r in rows:
            item = by_key.get(name_key(clean_text(r["component"]) or ""))
            n_old = len([e for e in r["entries"] if e["text"] or e["url"]])
            n_new = len(live_sources(item)) if item else 0
            widest = max(widest, n_old, n_new)

        ws = wb.create_sheet(p["name"][:31])
        header = (["Row", "Category", "Component", "Standard", "Qty", "Size"]
                  + ["Source %d" % (n + 1) for n in range(widest)]
                  + ["Notes", "Result"])
        ws.append(header)
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill, cell.font = HEAD_FILL, HEAD_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # The rebuild emits one row per component and build size. A component
        # listed at 250/300/350 keeps all three; the same component written
        # twice at the same size keeps the first. Vertically merged cells mean
        # size variants are not flagged as continuations, so grouping on
        # (component, size) is what actually decides.
        primary = {}
        for r in rows:
            comp = clean_text(r["component"])
            if not comp or is_drop_row(comp) or name_key(comp) in dropped:
                continue
            item = by_key.get(name_key(comp))
            if item is None:
                continue
            k = (item["id"], clean_size(r["size"]) or "")
            primary.setdefault(k, r["row"])

        counts = defaultdict(int)
        for r in rows:
            comp = clean_text(r["component"])
            out = ws.max_row + 1
            ws.cell(row=out, column=1, value=r["row"])

            old_links = [link_text(e["text"] or "", e["url"])
                         for e in r["entries"] if e["text"] or e["url"]]

            # --- rows that are not components -----------------------------
            if not comp or is_drop_row(comp) or name_key(comp) in dropped:
                kept = not (comp or "").strip().upper() == "OR"
                write_row(ws, out, widest,
                          [clean_text(r["category"]), comp, None, None, None],
                          old_links, [] if kept else None,
                          r["notes"], [],
                          "kept as a section marker" if kept else
                          "separator dropped: the alternatives it divided are "
                          "now one component",
                          gone=not kept)
                counts["structure kept" if kept else "separator dropped"] += 1
                continue

            item = by_key.get(name_key(comp))
            if item is None:
                write_row(ws, out, widest,
                          [clean_text(r["category"]), comp, None, None, None],
                          old_links, None, r["notes"], [],
                          "row removed", gone=True)
                counts["row removed"] += 1
                continue

            use = next((u for u in item["used_by"]
                        if u["project"] == p["id"]), None)
            pkey = (item["id"], clean_size(r["size"]) or "")
            kept_at = primary.get(pkey)
            merged_away = (use is not None and kept_at is not None
                           and r["row"] != kept_at)

            new_srcs = live_sources(item)
            new_links = ([link_text(label_for(s), s["url"]) for s in new_srcs]
                         + text_sources(item, p["id"]))
            new_notes = [n["text"] for n in item["notes"]
                         if p["id"] in n["projects"]]

            if merged_away:
                write_row(ws, out, widest,
                          [clean_text(r["category"]), comp, clean_text(r["standard"]),
                           clean_qty(r["qty"])[0], clean_size(r["size"])],
                          old_links, None, r["notes"], [],
                          'merged into row %d, "%s"'
                          % (kept_at, item["name"]), gone=True)
                counts["merged into another row"] += 1
                continue

            size = clean_size(r["size"])
            qty = clean_qty(r["qty"])[0]
            if use:
                match = next((q for q in use["qty"]
                              if (q.get("size") or None) == (size or None)), None)
                if match:
                    qty = match.get("qty")

            note = change_note(old_links, new_links, comp, item["name"])
            write_row(ws, out, widest,
                      [clean_category_pair(r["category"], item["category"]),
                       (comp, item["name"]),
                       (clean_text(r["standard"]), item["standard"]),
                       (clean_qty(r["qty"])[0], qty),
                       (size, size)],
                      old_links, new_links, r["notes"], new_notes, note)
            counts[note.split(",")[0] or "unchanged"] += 1

        widths = [7, 18, 40, 18, 8, 10] + [46] * widest + [50, 34]
        for n, w in enumerate(widths[:len(header)], start=1):
            ws.column_dimensions[get_column_letter(n)].width = w
        ws.freeze_panes = "C2"
        summary.append((p["name"], len(rows), dict(counts)))

    write_about(wb, summary)
    wb.save(OUT)
    print("wrote %s\n" % OUT)
    for name, n, counts in summary:
        head = ", ".join("%s: %d" % kv for kv in sorted(counts.items()))
        print("  %-22s %4d rows   %s" % (name, n, head))


def clean_category_pair(old, new):
    return (clean_text(old), new)


def write_row(ws, out, widest, meta, old_links, new_links, old_notes,
              new_notes, note, gone=False):
    """meta is five values, each either a plain value or an (old, new) pair."""
    for n, v in enumerate(meta):
        col = n + 2
        if isinstance(v, tuple):
            o, w = v
            cell_value = diff_cell(none_str(o), none_str(w))
        else:
            cell_value = none_str(v) or None
            if gone and cell_value:
                cell_value = CellRichText([TextBlock(DEL, cell_value)])
        ws.cell(row=out, column=col, value=cell_value)

    if new_links is None:                       # the whole row is going away
        pairs = [(o, "") for o in old_links]
    else:
        pairs = align_links(old_links, new_links)
    for n, (o, w) in enumerate(pairs[:widest]):
        c = ws.cell(row=out, column=7 + n, value=diff_link(o, w))
        c.alignment = WRAP
        if not o and w:
            c.fill = NEW_FILL
        elif o and not w:
            c.fill = GONE_FILL

    note_col = 7 + widest
    ws.cell(row=out, column=note_col,
            value=diff_cell(" / ".join(old_notes),
                            " / ".join(new_notes) if new_links is not None
                            else "")).alignment = WRAP
    nc = ws.cell(row=out, column=note_col + 1, value=note or None)
    nc.alignment = WRAP
    if gone:
        nc.font = Font(color="FF9B2C24")
        for c in range(1, note_col + 2):
            if not ws.cell(row=out, column=c).fill.fill_type:
                ws.cell(row=out, column=c).fill = GONE_FILL


def split_link(text):
    """-> (label, address). The address is display only."""
    m = re.search(r"^(.*?)\s*<([^>]+)>$", text or "")
    return (m.group(1), m.group(2)) if m else ((text or ""), "")


def diff_link(old, new):
    """Diff the vendor label; never redline the address itself.

    Links are paired on their canonical address, so a matched pair already
    points at the same product - showing `.../91290A136` struck through beside
    `.../91290A136/` added is churn, not a change. The rebuild's address is
    shown plainly and the "Result" column reports that addresses were cleaned.
    """
    if old and not new:
        return CellRichText([TextBlock(DEL, old)])
    if new and not old:
        return CellRichText([TextBlock(ADD, new)])
    if not old and not new:
        return None
    o_lab, _ = split_link(old)
    n_lab, n_url = split_link(new)
    tail = ("  <%s>" % n_url) if n_url else ""
    body = diff_cell(o_lab, n_lab)
    if body is None:
        return (n_lab + tail) or None
    if isinstance(body, str):
        return body + tail
    return CellRichText(list(body) + ([tail] if tail else []))


def none_str(v):
    return "" if v is None else str(v)


def url_of(text):
    """The address a rendered link cell points at, canonicalised.

    The rebuild cleans tracking junk and trailing slashes off every url, so
    comparing raw strings would report the same product as one link removed and
    another added. Matching on the canonical form keeps a tidied url as a small
    edit inside the cell instead.
    """
    m = re.search(r"<([^>]+)>$", text or "")
    if not m:
        return "text:" + (text or "").strip().lower()
    return url_key(m.group(1)) or m.group(1)


def align_links(old, new):
    """Pair links by url so a link that only changed column reads unchanged."""
    new_by_url = OrderedDict((url_of(x), x) for x in new)
    pairs, used = [], set()
    for o in old:
        u = url_of(o)
        if u in new_by_url:
            pairs.append((o, new_by_url[u]))
            used.add(u)
        else:
            pairs.append((o, ""))
    for u, w in new_by_url.items():
        if u not in used:
            pairs.append(("", w))
    return pairs


def change_note(old_links, new_links, old_name, new_name):
    bits = []
    new_keys, old_keys = {url_of(x) for x in new_links}, {url_of(x) for x in old_links}
    lost = [o for o in old_links
            if url_of(o) not in new_keys and not url_of(o).startswith("text:")]
    gained = [w for w in new_links
              if url_of(w) not in old_keys and not url_of(w).startswith("text:")]
    if lost and not new_links:
        bits.append("all links were dead")
    elif lost:
        bits.append("%d dead link%s removed" % (len(lost), "" if len(lost) == 1 else "s"))
    if gained:
        bits.append("%d link%s added from other tabs"
                    % (len(gained), "" if len(gained) == 1 else "s"))
    tidied = sum(1 for o, w in align_links(old_links, new_links)
                 if o and w and o != w)
    if tidied:
        bits.append("%d link address%s cleaned"
                    % (tidied, "" if tidied == 1 else "es"))
    if name_key(old_name) != name_key(new_name):
        bits.append("renamed to the shared spelling")
    return ", ".join(bits)


def write_about(wb, summary):
    ws = wb.create_sheet("About", 0)
    ws.append(["Voron sourcing guide - diff"])
    ws["A1"].font = Font(bold=True, size=14)
    for line in ["",
                 "The published guide against the rebuild, on the original's rows.",
                 "",
                 "Inside each cell:",
                 "    red struck through   removed",
                 "    blue bold            added",
                 "    plain                unchanged",
                 "",
                 "Links are matched on their address, so one that only moved "
                 "column shows as unchanged.",
                 "A row shaded pink is gone from the rebuild; the last column "
                 "says why.",
                 ""]:
        ws.append([line])
    ws.append(["Tab", "Original rows", "Result"])
    hdr = ws.max_row
    for c in range(1, 4):
        ws.cell(row=hdr, column=c).font = Font(bold=True)
    for name, n, counts in summary:
        ws.append([name, n, ", ".join("%s: %d" % kv for kv in sorted(counts.items()))])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 96
    ws["A1"].font = Font(bold=True, size=14)


if __name__ == "__main__":
    main()
